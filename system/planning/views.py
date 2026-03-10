from django.shortcuts import render, get_object_or_404, redirect
from .models import plannedDownTime, plannedDownTimeCells, plannedProduction, productionDetail
from system.core.models import Cell, Model
from django.views.generic import ListView, CreateView
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse
import calendar
from django.urls import reverse_lazy
from django.views.generic import CreateView
from .forms import PlannedProductionForm, ProductionDetailForm, plannedDownTimeForm, UploadExcelForm
from django.contrib import messages
import io
import openpyxl

# =========================================
# Lista de máquinas para producción
# =========================================
class machineListView(ListView):
    model = Cell
    template_name = 'planning/machineList.html'
    context_object_name = 'machines'
    
# =========================================
# Detalles de producción máquina 
# =========================================
def productionPlan(request, cell_id):
    cell = get_object_or_404(Cell, id=cell_id)
    week_offset = int(request.GET.get('week_offset', 0))
    today = datetime.now().date()
    monday = today - timedelta(days=today.weekday())
    target_monday = monday + timedelta(weeks=week_offset)
    workdays = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

    week_dates = []
    for i in range(5):  # 5 días laborales
        week_dates.append(target_monday + timedelta(days=i))
    
    # Obtener la producción planeada para esta celda y semana
    production_data = plannedProduction.objects.filter(
        cell=cell,
        date__range=[week_dates[0], week_dates[-1]]
    ).prefetch_related('details__model')
    
    # Organizar los datos por fecha
    production_by_date = {}
    for production in production_data:
        if production.date not in production_by_date:
            production_by_date[production.date] = []
        production_by_date[production.date].extend(production.details.all())
    
    # Preparar el contexto para el template
    calendar_data = []
    for date in week_dates:
        day_data = {
            'date': date,
            'day_name': date.strftime('%A'),
            'day_short': date.strftime('%a'),
            'day_number': date.day,
            'details': production_by_date.get(date, [])
        }
        calendar_data.append(day_data)
    
    # Calcular fechas para navegación
    prev_week_offset = week_offset - 1
    next_week_offset = week_offset + 1
    
    context = {
        'cell': cell,
        'calendar_data': calendar_data,
        'week_start': week_dates[0],
        'week_end': week_dates[-1],
        'workdays': workdays,
        'current_week_offset': week_offset,
        'prev_week_offset': prev_week_offset,
        'next_week_offset': next_week_offset,
        'week_range': f"{week_dates[0].strftime('%d/%m')} - {week_dates[-1].strftime('%d/%m/%Y')}"
    }

    return render(request, 'planning/plannedProduction.html', context)

# =========================================
# Lee el contenido del excel
# =========================================
def leer_excel(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    # Fecha en C2
    fecha = ws["C2"].value  

    # Headers de B4:E4
    headers = [ws.cell(row=4, column=col).value for col in range(2, 6)]

    # Datos desde fila 5
    data = []
    for row in ws.iter_rows(min_row=5, min_col=2, max_col=5, values_only=True):
        if all(v is None for v in row):
            continue
        data.append({
            "Celda": row[0],
            "WorkOrder": row[1],
            "Modelo": row[2],
            "Cantidad": row[3],
            "Fecha": fecha
        })

    return data

# =========================================
# Maneja los forms de agregar producción: preview, excel, manual
# =========================================
def addProduction(request):
    if request.method == "POST" and request.GET.get("preview") == "1":
        # ----------- PREVIEW VIA AJAX -----------
        excel_form = UploadExcelForm(request.POST, request.FILES)
        if excel_form.is_valid():
            file = excel_form.cleaned_data["file"]
            try:
                preview_data = leer_excel(file)

                # Validaciones simples
                errors = []
                for row in preview_data:
                    if not Cell.objects.filter(name=row["Celda"]).exists():
                        errors.append(f"Celda '{row['Celda']}' no existe.")
                    if not Model.objects.filter(name=row["Modelo"]).exists():
                        errors.append(f"Modelo '{row['Modelo']}' no existe.")
                    if row["Cantidad"] is None or row["Cantidad"] <= 0:
                        errors.append(f"Cantidad inválida en WorkOrder {row['WorkOrder']}.")

                return JsonResponse({
                    "preview": {
                        "headers": list(preview_data[0].keys()) if preview_data else [],
                        "rows": [list(r.values()) for r in preview_data]
                    },
                    "errors": errors
                })
            except Exception as e:
                return JsonResponse({"errors": [str(e)]})
        else:
            return JsonResponse({"errors": ["Archivo inválido."]})

    elif request.method == "POST" and "daily_submit" in request.POST:
        # ----------- GUARDAR EXCEL -----------
        excel_form = UploadExcelForm(request.POST, request.FILES)
        if excel_form.is_valid():
            file = excel_form.cleaned_data["file"]
            rows = leer_excel(file)

            for row in rows:
                cell = Cell.objects.get(name=row["Celda"])
                
                try:
                    model = Model.objects.get(name=row["Modelo"])
                except Model.DoesNotExist:
                    messages.error(request, f"Modelo '{row["Modelo"]}' no existe. Saltando WorkOrder {row["WorkOrder"]}.")
                    continue

                planned_production, created = plannedProduction.objects.get_or_create(
                    cell=cell,
                    date=row["Fecha"],
                    workorder=row["WorkOrder"],
                    defaults={'created_by': request.user}
                )

                production_detail, detail_created = productionDetail.objects.get_or_create(
                    planned_production=planned_production,
                    model=model,
                    defaults={'quantity': row["Cantidad"]}
                )
                if not detail_created:
                    production_detail.quantity += row["Cantidad"]
                    production_detail.save()

            messages.success(request, "Producción cargada exitosamente.")
            return redirect("planningMachineList")

    elif request.method == "POST":
        form = PlannedProductionForm(request.POST)
        detail_form = ProductionDetailForm(request.POST)
        if form.is_valid() and detail_form.is_valid():
            planned_production = form.save(commit=False)
            planned_production.created_by = request.user
            planned_production.save()

            production_detail = detail_form.save(commit=False)
            production_detail.planned_production = planned_production
            production_detail.save()

            messages.success(request, "Producción planeada agregada exitosamente.")
            return redirect("planningMachineList")
        else:
            messages.error(request, "Hubo un error al agregar la producción planeada.")

    # ----------- FORM NORMAL -----------
    form = PlannedProductionForm()
    detail_form = ProductionDetailForm()
    excel_form = UploadExcelForm()
    return render(request, "planning/addProduction.html", {
        "form": form,
        "detail_form": detail_form,
        "excel_form": excel_form
    })

# =========================================
# Lista de máquinas para downtime
# =========================================
class downTimeListView(ListView):
    model = Cell
    template_name = 'planning/downTimeList.html'
    context_object_name = 'machines'

# =========================================
# Detalles máquina tiempo muerto
# =========================================
def plannedDownTime (request, cell_id):
    cell = get_object_or_404(Cell, pk=cell_id)
    planned_downtime_cells = plannedDownTimeCells.objects.filter(cell=cell).select_related('planned_downtime')
    week_offset = int(request.GET.get('week_offset', 0))
    today = datetime.now().date()
    monday = today - timedelta(days=today.weekday())
    target_monday = monday + timedelta(weeks=week_offset)
    month = today.month
    month_name = calendar.month_name[month]
    year = today.year
    cal = calendar.monthcalendar(year, month)
    weekdays = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    workdays = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    start_time = datetime.strptime("06:30", "%H:%M")
    end_time = datetime.strptime("17:00", "%H:%M")
    interval = timedelta(minutes=90)  # 1 hora y media

    week_dates = []
    for i in range(5):  # 5 días laborales
        week_dates.append(target_monday + timedelta(days=i))

    time_slots = []
    current = start_time
    while current < end_time:
        # puedes guardar el inicio y fin de cada intervalo
        time_slots.append({
            "start": current.strftime("%H:%M"),
            "end": (current + interval).strftime("%H:%M")
        })
        current += interval
    
    # Calcular fechas para navegación
    prev_week_offset = week_offset - 1
    next_week_offset = week_offset + 1
     
    context = {
        'cells': cell,
        'downtimes': planned_downtime_cells,
        'today': today.day,
        'month': month_name, 
        'week_start': week_dates[0].strftime("%A %d"),
        'week_end': week_dates[-1].strftime("%A %d"),
        'month_days': cal,
        "weekdays": weekdays,
        'workdays': workdays,
        'year': year,
        'time_slots': time_slots
    }
    return render(request, 'planning/plannedDownTime.html', context)

# =========================================
# Maneja el form de agregar downtime
# =========================================
def addDownTime(request):
    if request.method == "POST":
        form = plannedDownTimeForm(request.POST)
        if form.is_valid():
            downtime = form.save(commit=False)
            downtime.created_by = request.user
            downtime.save()

            cells = form.cleaned_data["cells"]
            for cell in cells:
                plannedDownTimeCells.objects.create(
                    planned_downtime=downtime,
                    cell=cell
                )
            return redirect("downTimeList")
    else:
        form = plannedDownTimeForm()

    return render(request, 'planning/addDownTime.html', {
        "form": form,
    })
    
