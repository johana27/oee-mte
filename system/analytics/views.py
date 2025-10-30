from django.shortcuts import render, get_object_or_404
from django.views.generic import ListView, DetailView
from core.models import Cell, Cause
from .models import Recap
from manufacturing.models import Defect, DownTime
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Avg, Sum, Count, Q
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import redirect
# =========================================
# Definir roles
# =========================================
def es_lider(user):
    return user.groups.filter(name='lider').exists()

def es_ingeniero(user):
    return user.groups.filter(name__in=['admin', 'Ingeniero']).exists()

# =========================================
# Template por si no tiene permisos
# =========================================
def noAccess(request):
    return render(request, 'noAccess.html', status=403)

# =========================================
# Lista máquinas dashborads
# =========================================
class machineListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Cell
    template_name = 'analytics/analyticsList.html'
    context_object_name = 'machines'

    def test_func(self):
        return self.request.user.groups.filter(name__in=['admin', 'Ingeniero']).exists()
    
    def handle_no_permission(self):
        return redirect('noAccess')

# =========================================
# Obtiene el lunes y domingo
# =========================================
def weekRange(date=None):
    if date is None:
        date = timezone.now()

    date_only = date.date()

    days_since_monday = date_only.weekday()
    monday = date_only - timedelta(days=days_since_monday)
    start = datetime.combine(monday, datetime.min.time())
    start = timezone.make_aware(start)
    end = start + timedelta(days=6, hours=23, minutes=59, seconds=59)
    
    return start, end

# =========================================
# Dashboard máquina 
# =========================================
@login_required
@user_passes_test(es_ingeniero, login_url='noAccess')
def machineDashboard(request, cell_id):
    cell = get_object_or_404(Cell, id=cell_id)
    start, end = weekRange()
    
    # Obtener métricas usando el manager
    weekly = Recap.objects.weekly_metrics(cell_id, start, end)
    monthly = Recap.objects.monthly_metrics(cell_id)
    daily_oee = Recap.objects.daily_oee_data(cell_id)
    defects_list = Recap.objects.defectsList(cell_id)
    
    context = {
        # Métricas semanales
        "wk_availability": weekly["availability"],
        "wk_performance": weekly["performance"],
        "wk_quality": weekly["quality"],
        "week": weekly["oee"],
        "sum_planned_pieces": weekly["planned_pieces"],
        "sum_total_pieces": weekly["actual_pieces"],
        "completion_percentage": weekly["completion_percentage"],
        
        # Métricas mensuales
        "avg_availability": monthly["availability"],
        "avg_performance": monthly["performance"],
        "avg_quality": monthly["quality"],
        "month": monthly["oee"],
        "sum_downtime": monthly["downtime"],
        "operating_minutes": monthly["operating_minutes"],
        "total_defects": monthly["sum_total_defects"],
        "defects_list": defects_list,
        
        "chart_labels": daily_oee['labels'],
        "chart_data": daily_oee['data'],
        "chart_colors": daily_oee['colors'],
        "chart_month": daily_oee['month_name'],

        # Otros
        "cell": cell,
    }
    
    return render(request, 'analytics/machineDashboard.html', context)

# =========================================
# Dashboard de todo el piso
# =========================================
@login_required
@user_passes_test(es_ingeniero, login_url='noAccess')
def plantDashboard(request):
    now = timezone.now()
    
    # Obtener rango de la semana actual
    week_start, week_end = weekRange()
    
    # Obtener todas las celdas activas
    cells = Cell.objects.filter(is_active=True)
    
    # ===== 1. RESUMEN DE PRODUCCIÓN SEMANAL =====
    weekly_summary = get_weekly_summary(week_start, week_end)
    
    # ===== 2. OEE MENSUAL GENERAL =====
    monthly_oee = get_monthly_oee()
    
    # ===== 3. GRÁFICA OEE DIARIO DEL MES (TODAS LAS CELDAS) =====
    daily_oee_chart = get_daily_oee_chart(now.year, now.month)
    
    # ===== 4. CELDAS CON MAYOR Y MENOR OEE DE LA SEMANA =====
    cells_ranking = get_cells_weekly_ranking(cells, week_start, week_end)
    
    # ===== 5. PRINCIPALES CAUSAS DE DEFECTOS DEL MES =====
    top_defects = get_top_defects(now.year, now.month, limit=10)
    
    # ===== 6. PRINCIPALES CAUSAS DE TIEMPOS MUERTOS DEL MES =====
    top_downtimes = get_top_downtimes(now.year, now.month, limit=10)
    
    context = {
        # Resumen semanal
        'weekly_planned': weekly_summary['planned_pieces'],
        'weekly_actual': weekly_summary['actual_pieces'],
        'weekly_effectiveness': weekly_summary['effectiveness'],
        'week_start': week_start,
        'week_end': week_end,
        
        # OEE mensual
        'monthly_oee': monthly_oee['oee'],
        'monthly_availability': monthly_oee['availability'],
        'monthly_performance': monthly_oee['performance'],
        'monthly_quality': monthly_oee['quality'],
        
        # Gráfica OEE diario
        'daily_labels': daily_oee_chart['labels'],
        'daily_data': daily_oee_chart['data'],
        'daily_colors': daily_oee_chart['colors'],
        'chart_month': daily_oee_chart['month_name'],
        
        # Ranking de celdas
        'best_cells': cells_ranking['best'][:3],
        'worst_cells': cells_ranking['worst'][:3],
        
        # Top defectos
        'top_defects': top_defects,
        'total_defects': sum(d['total'] for d in top_defects),
        
        # Top tiempos muertos
        'top_downtimes': top_downtimes,
        'total_downtime_minutes': sum(d['total_minutes'] for d in top_downtimes),
        
        # Información general
        'current_date': now,
        'current_month': now.strftime('%B %Y'),
    }
    
    return render(request, 'analytics/plantDashboard.html', context)


# =========================================
# FUNCIONES AUXILIARES
# =========================================

def get_weekly_summary(start_date, end_date):
    """Obtiene resumen de producción semanal de todas las celdas"""
    recaps = Recap.objects.filter(
        pub_date__gte=start_date,
        pub_date__lte=end_date
    )
    
    if not recaps.exists():
        return {
            'planned_pieces': 0,
            'actual_pieces': 0,
            'effectiveness': 0,
        }
    
    summary = recaps.aggregate(
        planned=Sum('total_planned_pieces'),
        actual=Sum('total_actual_pieces')
    )
    
    planned = summary['planned'] or 0
    actual = summary['actual'] or 0
    effectiveness = (actual * 100 / planned) if planned > 0 else 0
    
    return {
        'planned_pieces': planned,
        'actual_pieces': actual,
        'effectiveness': round(effectiveness, 1),
    }


def get_monthly_oee():
    """Calcula OEE mensual promedio de todas las celdas"""
    now = timezone.now()
    
    recaps = Recap.objects.filter(
        pub_date__year=now.year,
        pub_date__month=now.month
    )
    
    if not recaps.exists():
        return {
            'oee': 0,
            'availability': 0,
            'performance': 0,
            'quality': 0,
        }
    
    metrics = recaps.aggregate(
        avg_availability=Avg('availability'),
        avg_performance=Avg('performance'),
        avg_quality=Avg('quality'),
    )
    
    availability = metrics['avg_availability'] or 0
    performance = metrics['avg_performance'] or 0
    quality = metrics['avg_quality'] or 0
    oee = availability * performance * quality * 100
    
    return {
        'oee': round(oee, 2),
        'availability': round(availability * 100, 2),
        'performance': round(performance * 100, 2),
        'quality': round(quality * 100, 2),
    }


def get_daily_oee_chart(year, month):
    """Obtiene OEE promedio diario de todas las celdas para el mes"""
    num_days = monthrange(year, month)[1]
    
    labels = []
    data = []
    colors = []
    
    for day in range(1, num_days + 1):
        labels.append(str(day))
        
        # Obtener todos los recaps de ese día
        day_recaps = Recap.objects.filter(
            pub_date__year=year,
            pub_date__month=month,
            pub_date__day=day
        )
        
        if day_recaps.exists():
            # Calcular OEE promedio del día
            metrics = day_recaps.aggregate(
                avg_availability=Avg('availability'),
                avg_performance=Avg('performance'),
                avg_quality=Avg('quality'),
            )
            
            availability = metrics['avg_availability'] or 0
            performance = metrics['avg_performance'] or 0
            quality = metrics['avg_quality'] or 0
            oee = availability * performance * quality * 100
            
            data.append(round(oee, 2))
            
            # Color según el OEE
            if oee >= 85:
                colors.append('rgba(34, 197, 94, 0.8)')  # Verde
            elif oee >= 70:
                colors.append('rgba(251, 191, 36, 0.8)')  # Amarillo
            else:
                colors.append('rgba(239, 68, 68, 0.8)')  # Rojo
        else:
            # Sin datos
            data.append(0)
            colors.append('rgba(156, 163, 175, 0.3)')  # Gris
    
    return {
        'labels': labels,
        'data': data,
        'colors': colors,
        'month_name': datetime(year, month, 1).strftime('%B %Y')
    }


def get_cells_weekly_ranking(cells, start_date, end_date):
    """Obtiene ranking de celdas por OEE semanal"""
    cells_data = []
    
    for cell in cells:
        recaps = Recap.objects.filter(
            cell=cell,
            pub_date__gte=start_date,
            pub_date__lte=end_date
        )
        
        if not recaps.exists():
            continue
        
        metrics = recaps.aggregate(
            avg_availability=Avg('availability'),
            avg_performance=Avg('performance'),
            avg_quality=Avg('quality'),
        )
        
        availability = metrics['avg_availability'] or 0
        performance = metrics['avg_performance'] or 0
        quality = metrics['avg_quality'] or 0
        oee = availability * performance * quality * 100
        
        cells_data.append({
            'cell': cell,
            'oee': round(oee, 2),
            'availability': round(availability * 100, 2),
            'performance': round(performance * 100, 2),
            'quality': round(quality * 100, 2),
        })
    
    # Ordenar por OEE
    cells_data.sort(key=lambda x: x['oee'], reverse=True)
    
    return {
        'best': cells_data,  # Las mejores están al inicio
        'worst': list(reversed(cells_data)),  # Las peores al invertir
    }


def get_top_defects(year, month, limit=10):
    """Obtiene principales causas de defectos del mes"""
    defects = (
        Defect.objects.filter(
            created_at__year=year,
            created_at__month=month,
            cause__type='defect'
        )
        .values('cause__name')
        .annotate(
            total=Sum('quantity'),
            occurrences=Count('id')
        )
        .order_by('-total')[:limit]
    )
    
    return [
        {
            'cause': d['cause__name'],
            'total': d['total'],
            'occurrences': d['occurrences'],
        }
        for d in defects
    ]


def get_top_downtimes(year, month, limit=10):
    """Obtiene principales causas de tiempos muertos del mes"""
    downtimes = DownTime.objects.filter(
        created_at__year=year,
        created_at__month=month,
        cause__type='downtime',
        start__isnull=False,
        end__isnull=False
    )
    
    # Agrupar por causa y calcular duración total
    cause_data = {}
    
    for dt in downtimes:
        cause_name = dt.cause.name if dt.cause else 'Sin causa'
        duration = dt.duration_minutes
        
        if cause_name not in cause_data:
            cause_data[cause_name] = {
                'total_minutes': 0,
                'occurrences': 0
            }
        
        cause_data[cause_name]['total_minutes'] += duration
        cause_data[cause_name]['occurrences'] += 1
    
    # Convertir a lista y ordenar
    result = [
        {
            'cause': cause,
            'total_minutes': round(data['total_minutes'], 1),
            'total_hours': round(data['total_minutes'] / 60, 1),
            'occurrences': data['occurrences'],
        }
        for cause, data in cause_data.items()
    ]
    
    result.sort(key=lambda x: x['total_minutes'], reverse=True)
    
    return result[:limit]

# =========================================
# Vista para reportes
# =========================================
@login_required
@user_passes_test(es_ingeniero, login_url='noAccess')
def Reports(request):
    # Si se solicita generar el reporte
    if request.GET.get('generate'):
        # Obtener parámetros
        report_type = request.GET.get('report_type', 'monthly')
        cell_id = request.GET.get('cell_id')
        year = int(request.GET.get('year', timezone.now().year))
        month = int(request.GET.get('month', timezone.now().month)) if report_type in ['weekly', 'monthly', 'quarterly'] else None
        week = int(request.GET.get('week', 1)) if report_type == 'weekly' else None
        
        # Validar que se seleccionó una celda
        if not cell_id:
            return HttpResponse("Debe seleccionar una celda", status=400)
        
        try:
            cell = Cell.objects.get(id=cell_id, is_active=True)
        except Cell.DoesNotExist:
            return HttpResponse("Celda no encontrada", status=404)
        
        # Calcular rango de fechas
        if report_type == 'weekly':
            first_day = datetime(year, month, 1)
            start_date = first_day + timedelta(weeks=week-1)
            end_date = start_date + timedelta(days=6)
            period_label = f"Semana {week}, {datetime(year, month, 1).strftime('%B %Y')}"
        elif report_type == 'monthly':
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            period_label = datetime(year, month, 1).strftime('%B %Y')
        elif report_type == 'quarterly':
            # 3 meses atrás desde el mes seleccionado
            start_date = datetime(year, month, 1) - timedelta(days=90)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            period_label = f"Últimos 3 meses hasta {datetime(year, month, 1).strftime('%B %Y')}"
        elif report_type == 'semiannual':
            # 6 meses - primer o segundo semestre
            if month <= 6:
                start_date = datetime(year, 1, 1)
                end_date = datetime(year, 6, 30)
                period_label = f"Primer Semestre {year}"
            else:
                start_date = datetime(year, 7, 1)
                end_date = datetime(year, 12, 31)
                period_label = f"Segundo Semestre {year}"
        else:  # annual
            start_date = datetime(year, 1, 1)
            end_date = datetime(year, 12, 31)
            period_label = f"Año {year}"
        
        # Hacer fechas timezone-aware
        start_date = timezone.make_aware(start_date)
        end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Producción"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Headers
        headers = [
            'Fecha', 'Celda', 'Tipo_Registro', 
            # Campos Recap
            'Piezas_Planeadas', 'Piezas_Producidas', 'Total_Defectos', 
            'Tiempo_Muerto_Min', 'Minutos_Operacion', 'Disponibilidad_%', 
            'Rendimiento_%', 'Calidad_%', 'OEE_%',
            # Campos Downtime
            'Causa_Downtime', 'Inicio_Downtime', 'Fin_Downtime', 
            'Duracion_Downtime_Min', 'Comentarios_Downtime',
            # Campos Defecto
            'Causa_Defecto', 'Cantidad_Defecto', 'Tipo_Defecto', 
            'Comentarios_Defecto', 'Modelo_Defecto'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell_obj = ws.cell(row=1, column=col)
            cell_obj.value = header
            cell_obj.fill = header_fill
            cell_obj.font = header_font
            cell_obj.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        
        # ============ RECAPS DIARIOS ============
        recaps = Recap.objects.filter(
            cell=cell,
            pub_date__gte=start_date,
            pub_date__lte=end_date
        ).select_related('cell').order_by('pub_date')
        
        for recap in recaps:
            ws.cell(row=row, column=1).value = recap.pub_date.date()
            ws.cell(row=row, column=2).value = cell.name
            ws.cell(row=row, column=3).value = 'RECAP'
            ws.cell(row=row, column=4).value = recap.total_planned_pieces
            ws.cell(row=row, column=5).value = recap.total_actual_pieces
            ws.cell(row=row, column=6).value = recap.total_defects
            ws.cell(row=row, column=7).value = recap.total_downtime_minutes
            ws.cell(row=row, column=8).value = recap.total_operating_minutes
            ws.cell(row=row, column=9).value = round(recap.availability * 100, 2)
            ws.cell(row=row, column=10).value = round(recap.performance * 100, 2)
            ws.cell(row=row, column=11).value = round(recap.quality * 100, 2)
            ws.cell(row=row, column=12).value = round(recap.oee_percentage, 2)
            row += 1
        
        # ============ DOWNTIMES ============
        downtimes = DownTime.objects.filter(
            cell=cell,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).select_related('cell', 'cause', 'created_by').order_by('created_at')

        for downtime in downtimes:
            ws.cell(row=row, column=1).value = downtime.created_at.date()
            ws.cell(row=row, column=2).value = cell.name
            ws.cell(row=row, column=3).value = 'DOWNTIME'
            ws.cell(row=row, column=13).value = downtime.cause.name if downtime.cause else ''
            # Convert timezone-aware datetimes to naive datetimes
            ws.cell(row=row, column=14).value = downtime.start.replace(tzinfo=None) if downtime.start else None
            ws.cell(row=row, column=15).value = downtime.end.replace(tzinfo=None) if downtime.end else None
            ws.cell(row=row, column=16).value = downtime.duration_minutes
            ws.cell(row=row, column=17).value = downtime.comments or ''
            row += 1
        
        # ============ DEFECTOS ============
        defects = Defect.objects.filter(
            production_detail__planned_production__cell=cell,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).select_related(
            'cause', 'production_detail__model', 'created_by'
        ).order_by('created_at')
        
        for defect in defects:
            ws.cell(row=row, column=1).value = defect.created_at.date()
            ws.cell(row=row, column=2).value = cell.name
            ws.cell(row=row, column=3).value = 'DEFECTO'
            ws.cell(row=row, column=18).value = defect.cause.name if defect.cause else ''
            ws.cell(row=row, column=19).value = defect.quantity
            ws.cell(row=row, column=20).value = defect.type or ''
            ws.cell(row=row, column=21).value = defect.comments or ''
            ws.cell(row=row, column=22).value = defect.model.name if defect.model else ''
            row += 1
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 12, 'B': 15, 'C': 15, 'D': 16, 'E': 18, 'F': 14,
            'G': 18, 'H': 18, 'I': 16, 'J': 14, 'K': 12, 'L': 10,
            'M': 20, 'N': 18, 'O': 18, 'P': 20, 'Q': 30,
            'R': 20, 'S': 16, 'T': 14, 'U': 30, 'V': 18
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"datos_produccion_{cell.name}_{report_type}_{year}"
        if month:
            filename += f"_{month:02d}"
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb.save(response)
        return response
    
    # Mostrar formulario
    cells = Cell.objects.filter(is_active=True).order_by('name')
    current_year = timezone.now().year
    current_month = timezone.now().month
    
    # Generar lista de años (últimos 5 años)
    years = range(current_year - 4, current_year + 1)
    
    # Lista de meses
    months = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    # Lista de semanas
    weeks = range(1, 5)
    
    context = {
        'cells': cells,
        'years': years,
        'months': months,
        'weeks': weeks,
        'current_year': current_year,
        'current_month': current_month,
    }
    
    return render(request, 'analytics/reports.html', context)